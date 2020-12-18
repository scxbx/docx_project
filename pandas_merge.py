import winreg
from collections import Counter

import pandas as pd
import os
import openpyxl
from openpyxl.styles import Alignment


def findDuplicatedElements(mylist):
    b = dict(Counter(mylist))
    return [key for key, value in b.items() if value > 1]  # 只展示重复元素
    # print({key: value for key, value in b.items() if value > 1})  # 展现重复元素和重复次数


def get_desktop():
    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                         r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders')
    return winreg.QueryValueEx(key, "Desktop")[0]


def merge_a_family(in_filename, out_filename, population):
    wb = openpyxl.load_workbook(in_filename)
    ws = wb.active

    nrow = ws.max_row
    print(nrow)

    merge_col_num_list = [1, 2, 3, 9, 10]
    current_merge_head = 2
    family_count = 0
    for i in range(1 + 1, nrow + 1):
        if ws.cell(i, 1).value is not None:
            if i != 2 and i - 1 > current_merge_head:
                for j in merge_col_num_list:
                    ws.merge_cells(start_row=current_merge_head, start_column=j, end_row=i - 1, end_column=j)
                    ws.cell(current_merge_head, j).alignment = Alignment(vertical='center')

            if i != 2 and i - 1 - current_merge_head + 1 != ws.cell(current_merge_head, 3).value:
                print('家庭成员数量错误！ 户主：{}'.format(ws.cell(current_merge_head, 2).value))
            current_merge_head = i
            family_count += 1
            ws.cell(current_merge_head, 1).value = family_count
        if i == nrow:
            for j in merge_col_num_list:
                ws.merge_cells(start_row=current_merge_head, start_column=j, end_row=i, end_column=j)
                ws.cell(current_merge_head, j).alignment = Alignment(vertical='center')
    ws.append(['合计', family_count, population])

    # 行宽 列高
    ws.column_dimensions['A'].width = 5.78
    ws.column_dimensions['B'].width = 6.56
    ws.column_dimensions['C'].width = 6.89
    ws.column_dimensions['D'].width = 9.67
    ws.column_dimensions['E'].width = 7.22
    ws.column_dimensions['F'].width = 5.56
    ws.column_dimensions['G'].width = 23.89
    ws.column_dimensions['H'].width = 9.11
    ws.column_dimensions['I'].width = 38.44
    ws.column_dimensions['J'].width = 14.44
    ws.column_dimensions['K'].width = 21.33

    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 16.1


    wb.save(out_filename)


row_to_start = int(input('请输入数据开始的行数：'))
drop_index_list = []
# 汇总表从第 n 行开始有数据 出去表头还要drop n-2 行
for i in range(0, row_to_start - 2):
    drop_index_list.append(i)

# 文件路径
file_dir = os.path.join(get_desktop(), 'to_merge')

# 构建新的表格名称
new_filename = os.path.join(get_desktop(), '大汇总.xlsx')
# 找到文件路径下的所有表格名称，返回列表
file_list = os.listdir(file_dir)
new_list = []

for file in file_list:
    # 重构文件路径
    file_path = os.path.join(file_dir, file)
    # 将excel转换成DataFrame
    dataframe = pd.read_excel(file_path)

    '''
    for i in range(10):
        for j in range(10):
            print('{}, {}: {}'.format(i, j, dataframe.iat[i, j]))
    '''

    dataframe.drop([len(dataframe) - 1], inplace=True)
    dataframe.drop(index=drop_index_list, inplace=True)

    # print(dataframe)

    # 保存到新列表中
    new_list.append(dataframe)

# 多个DataFrame合并为一个
df = pd.concat(new_list)

# 写入到一个新excel表中
df.to_excel(new_filename, index=False)
print('总计：', len(df), '人')

# for i in range(len(df)):
# print(df.iloc[0: len(df), 6])
my_list = df.iloc[0: len(df), 6].tolist()
# print(my_list)
duplicatedList = findDuplicatedElements(my_list)
print('以下身份证重复：')

dup_list_filename = os.path.join(get_desktop(), '重复身份证.txt')
f = open(dup_list_filename, 'w')
for duplicatedItem in duplicatedList:
    print(duplicatedItem)
    f.write(str(duplicatedItem))
    f.write('\n')
f.close()

merge_a_family(new_filename, new_filename, len(df))

print('已生成文件：{} {}'.format(new_filename, dup_list_filename))

input('Press enter to quit program.')
