import os

import openpyxl
import win32com.client


def xls_to_xlsx(folder_path, file_name):
    """
    excel  .xls 后缀 改成 .xlsx 后缀
    folder_path 文件夹路径
    file_name 文件名字 带后缀 比如 aa.xls
    """
    folder_path = folder_path.replace('/', '\\')
    file_name = file_name.replace('/', '\\')
    name, suffix = file_name.split('.')
    excel_file_path = os.path.join(folder_path, file_name)

    excel = win32com.client.DispatchEx('Excel.Application')  # 要看MIME手册
    wb = excel.Workbooks.Open(excel_file_path)
    suffix = f".{suffix}x"
    new_file_name = f"{name}{suffix}"
    new_excel_file_path = os.sep.join([folder_path, new_file_name])
    # tset
    # print("new_excel_file_path: " + new_excel_file_path)
    wb.SaveAs(new_excel_file_path, FileFormat=51)
    wb.Close()
    excel.Application.Quit()

    os.remove(excel_file_path)
    return new_excel_file_path


def get_last_line_of_summary(full_filename, capital, all_population):
    wb = openpyxl.load_workbook(full_filename, data_only=True)
    ws = wb.active
    filename = os.path.split(full_filename)[1]
    for i in range(ws.max_row):
        if ws.cell(i + 1, 1).value == '合计':
            num_family = int(ws.cell(i + 1, 2).value)
            population = int(ws.cell(i + 1, 3).value)
            try:
                capital_per_group = capital / all_population * population
                print('{}\t{}\t{}\t{}\t{}'.format(filename, num_family, population, population * 10, capital_per_group))

            except:
                print(filename)


if __name__ == '__main__':
    capital = float(input('请输入本村委会的经营性资产：'))
    all_population = int(input('请输入本村委会的总人数：'))
    path = r'C:\Users\sc\Desktop\others\定安县新竹镇\新竹镇汇总结果\各村小组汇总表\祖坡村确认汇总表打印'

    for filename_in in os.listdir(path):
        # print(os.path.join(path, filename))
        # print(filename_in)
        old_suffix = filename_in.split('.')[-1]

        if old_suffix == 'xls':
            # print("transform .xls to .xlsx")
            excel_path = xls_to_xlsx(path, filename_in)
            # print(excel_path)
        elif old_suffix == 'xlsx':
            # print("no need to transform file type")
            excel_path = os.sep.join([path, filename_in])
            # print(excel_path)
            # print("wrong file type: " + old_suffix)
        else:
            print('wrong type!')
            excel_path = ''
        get_last_line_of_summary(excel_path, capital, all_population)

    # get_last_line_of_summary(os.path.join(r'C:\Users\sc\Desktop\others\定安县新竹镇\白堆村委会\白堆打印', '白堆村白堆二村民小组-36户-已校对_汇总表.xlsx'))
    input('Press any key to quit program.')
