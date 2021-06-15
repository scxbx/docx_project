import os
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, colors


def all_in_one(filename_in):
    folder_path, file_name = os.path.split(filename_in)
    book_in = openpyxl.load_workbook(filename_in, data_only=True)

    sheets_in = book_in.sheetnames

    sheet_in_2 = book_in[sheets_in[1]]
    sheet_in_4 = book_in[sheets_in[3]]
    sheet_in_2.delete_rows(23)
    sheet_in_2.delete_rows(22)

    if len(sheets_in) > 4:
        for i in range(len(sheets_in) - 4):
            sheet_in_i = book_in[sheets_in[i + 4]]
            for j in range(54):
                if sheet_in_i.cell(j + 4, 3).value is not None:
                    sheet_in_4.cell(54 * i + j + 59, 1).value = 54 * i + j + 56
                    sheet_in_4.cell(54 * i + j + 59, 3).value = sheet_in_i.cell(j + 4, 3).value
                    sheet_in_4.cell(54 * i + j + 59, 4).value = sheet_in_i.cell(j + 4, 4).value

                    sheet_in_4.cell(54 * i + j + 59, 1).alignment = Alignment(vertical='center', horizontal='center')
                    sheet_in_4.cell(54 * i + j + 59, 3).alignment = Alignment(wrapText=True, vertical='center',
                                                                              horizontal='center')
                    sheet_in_4.cell(54 * i + j + 59, 4).alignment = Alignment(wrapText=True)

            book_in.remove(sheet_in_i)

    row_count = 0
    for row in sheet_in_4:
        if not all([cell.value == None for cell in row]):
            row_count += 1

    for i in range(4, row_count + 1):
        sheet_in_4.row_dimensions[i].height = 25

    sheet_in_4.column_dimensions['A'].width = 4.0
    sheet_in_4.column_dimensions['F'].width = 8.0

    border_set = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    for row in sheet_in_4.iter_rows(min_row=4, max_col=7, max_row=row_count):
        for cell in row:
            cell.font = Font(size=9)
            cell.border = border_set

    filename_save = (os.path.split(filename_in))[1].split('.')[0] + '改.xlsx'
    folder_path2 = folder_path + r'\..\new menu'
    whole_save = os.path.join(folder_path2, filename_save)
    book_in.save(whole_save)

    return


if __name__ == '__main__':

    print(os.getcwd())
    path = os.path.join(os.getcwd(), 'old menu')
    new_path = os.path.join(os.getcwd(), 'new menu')
    for filename_in in os.listdir(path):
        print(os.path.join(path, filename_in))
        # read_summary(os.path.join(path, filename_in))
        all_in_one(os.path.join(path, filename_in))
    input('Press any key to quit program.')
