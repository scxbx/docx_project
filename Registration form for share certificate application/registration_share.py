import os

import openpyxl
import win32com
from openpyxl.styles import Alignment, Border, Side


def xls_to_xlsx(folder_path, file_name):
    """
    excel  .xls 后缀 改成 .xlsx 后缀
    folder_path 文件夹路径
    file_name 文件名字 带后缀 比如 aa.xls
    """
    folder_path = folder_path.replace('/', '\\')
    file_name = file_name.replace('/', '\\')
    name, suffix = file_name.split('.')
    excel_file_path = os.path.join(folder_path, file_name)

    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')  # 要看MIME手册
    wb = excel.Workbooks.Open(excel_file_path)
    suffix = f".{suffix}x"
    new_file_name = f"{name}{suffix}"
    new_excel_file_path = os.sep.join([folder_path, new_file_name])
    # tset
    print("new_excel_file_path: " + new_excel_file_path)
    wb.SaveAs(new_excel_file_path, FileFormat=51)
    wb.Close()
    excel.Application.Quit()
    return new_excel_file_path

def gen_registration_share(sample_path, summary_path):
    print('-----------------------------------------开始----------------------------------------------')
    print("模板: " + sample_path)
    print('输入汇总表：' + summary_path)
    folder_path, file_name = os.path.split(sample_path)
    # folder_path = r"C:\Users\sc\Desktop\test_openpyxl"
    # file_name = 'sample.xlsx'
    old_suffix = file_name.split('.')[-1]
    if old_suffix == 'xls':
        print("transform .xls to .xlsx")
        excel_path = xls_to_xlsx(folder_path, file_name)
    elif old_suffix == 'xlsx':
        print("no need to transform file type")
        excel_path = os.sep.join([folder_path, file_name])
        # print(excel_path)
    else:
        print("wrong file type: " + old_suffix)
        return

    folder_path2, file_name2 = os.path.split(summary_path)
    # folder_path = r"C:\Users\sc\Desktop\test_openpyxl"
    # file_name = 'sample.xlsx'
    old_suffix2 = file_name2.split('.')[-1]
    if old_suffix2 == 'xls':
        print("transform .xls to .xlsx")
        excel_path2 = xls_to_xlsx(folder_path2, file_name2)
    elif old_suffix2 == 'xlsx':
        print("no need to transform file type")
        excel_path2 = os.sep.join([folder_path2, file_name2])
        # print(excel_path)
    else:
        print("wrong file type: " + old_suffix2)
        return

    book_summary = openpyxl.load_workbook(excel_path2)
    sheet_summary = book_summary.active

    book_sample = openpyxl.load_workbook(excel_path)
    sheet_sample = book_sample.active

    # for data in sheet_in['A']:
    #    print(data.value)

    # Get the number of rows in worksheet
    max_row = len([row for row in sheet_summary if not all([cell.value is None for cell in row])])
    # print(max_row)

    sample_row = 3
    for row in range(3, max_row):
        serial_number = sheet_summary.cell(row + 1, 1).value
        if serial_number is not None:
            sample_row += 1
            # print(serial_number)
            sheet_sample.cell(sample_row, 1).value = serial_number
            sheet_sample.cell(sample_row, 3).value = sheet_summary.cell(row + 1, 2).value
            sheet_sample.cell(sample_row, 4).value = sheet_summary.cell(row + 1, 7).value
            sheet_sample.cell(sample_row, 5).value = sheet_summary.cell(row + 1, 10).value
            sheet_sample.cell(sample_row, 6).value = sheet_summary.cell(row + 1, 4).value
            sheet_sample.cell(sample_row, 7).value = sheet_summary.cell(row + 1, 3).value
            sheet_sample.cell(sample_row, 8).value = int(sheet_summary.cell(row + 1, 3).value) * 10

        else:
            # print('1. {}'.format(sheet_sample.cell(sample_row, 6).value))
            # print('2. {}'.format(sheet_summary.cell(row + 1, 4).value))
            if sheet_sample.cell(sample_row, 6).value is None:
                sheet_sample.cell(sample_row, 6).value = ''
            sheet_sample.cell(sample_row, 6).value += ('、' + sheet_summary.cell(row + 1, 4).value)
            # print('3. {}'.format(sheet_sample.cell(sample_row, 6).value))

        if serial_number in ['合计', '总计']:
            sheet_sample.cell(sample_row, 1).value = '合计'
            # print('row {}'.format(sample_row))
            sheet_sample.cell(sample_row, 1).alignment = Alignment(horizontal='center', vertical='center')
            sheet_sample.merge_cells(start_row=sample_row, start_column=1, end_row=sample_row, end_column=2)
            for i in [3, 4, 5, 6, 10, 11, 12]:
                sheet_sample.cell(sample_row, i).value = '―'

        # Format
        sheet_sample.row_dimensions[sample_row].height = 50.1
        for my_col in range(12):
            if my_col == 5:
                sheet_sample.cell(sample_row, my_col + 1).alignment \
                    = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                sheet_sample.cell(sample_row, my_col + 1).alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            sheet_sample.cell(sample_row, my_col + 1).border = thin_border

    for i in range(12):
        if sheet_summary.cell(2, i + 1).value is not None \
                and '股份' in sheet_summary.cell(2, i + 1).value \
                and ('：' in sheet_summary.cell(2, i + 1).value or ':' in sheet_summary.cell(2, i + 1).value):

            second_half = sheet_summary.cell(2, i + 1).value.split('：')[1] \
                if '：' in sheet_summary.cell(2, i + 1).value \
                else sheet_summary.cell(2, i + 1).value.split(':')[1]
            print('组织名称：{}'.format(second_half))
            sheet_sample.cell(2, 1).value = '组织名称：{}'.format(second_half)
    filename_save = (os.path.split(summary_path))[1].split('.')[0] + '股权证申领登记表.xlsx'
    folder_path2 = folder_path + r'\..\result'
    whole_save = os.path.join(folder_path2, filename_save)
    book_sample.save(whole_save)
    print('生成文件{}\n'.format(whole_save))


if __name__ == '__main__':
    path = os.path.join(os.getcwd(), 'summary')
    sample_f = os.path.join(os.getcwd(), 'sample')
    for filename_in in os.listdir(path):
        # read_summary(os.path.join(path, filename_in))
        gen_registration_share(os.path.join(sample_f, 'sample.xlsx'), os.path.join(path, filename_in))
    input('生成完毕。按回车键以结束程序。')
