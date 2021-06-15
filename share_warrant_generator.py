import os
from pathlib import Path

import win32com.client
import openpyxl
from tkinter import *
from tkinter import filedialog


def xls_to_xlsx(folder_path, file_name):
    """
    excel  .xls 后缀 改成 .xlsx 后缀
    folder_path 文件夹路径
    file_name 文件名字 带后缀 比如 aa.xls
    """
    folder_path = folder_path.replace('/', '\\')
    file_name = file_name.replace('/', '\\')
    name, suffix = file_name.split('.')
    excel_file_path = os.path.join(folder_path, file_name)

    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')  # 要看MIME手册
    wb = excel.Workbooks.Open(excel_file_path)
    suffix = f".{suffix}x"
    new_file_name = f"{name}{suffix}"
    new_excel_file_path = os.sep.join([folder_path, new_file_name])
    # test
    print("new_excel_file_path: " + new_excel_file_path)
    wb.SaveAs(new_excel_file_path, FileFormat=51)
    wb.Close()
    excel.Application.Quit()
    return new_excel_file_path


def generate_one_warrant(excel_path, confirm_sheet, index, para_list):
    warrant_book = openpyxl.load_workbook(excel_path, data_only=True)
    warrant_sheet1 = warrant_book["1"]  # 股权证首页
    warrant_sheet2 = warrant_book["2"]  # 股东基本信息登记
    warrant_sheet3 = warrant_book["3"]  # 股权登记
    # warrant_sheet2['E25'] = confirm_sheet['A10'].value

    confirm_row_index = 10
    warrant_sheet2_row_index = 25
    warrant_sheet3_row_index = 24

    sheet_id = confirm_sheet['K2'].value

    count = 0
    while confirm_sheet.cell(confirm_row_index, 1).value is not None \
            and not (('家庭' in confirm_sheet.cell(confirm_row_index, 1).value)
                     and (confirm_sheet.cell(confirm_row_index, 3).value is None)):
        member_name = confirm_sheet.cell(confirm_row_index, 1).value
        id_number = confirm_sheet.cell(confirm_row_index, 5).value
        gender = ''
        if id_number is None or id_number == '':
            print("警告：编号为{}的表中身份证号缺失".format(sheet_id))
        else:
            id_number = id_number.strip()
            if len(id_number) != 18:
                print("警告：编号为{}的表中身份证号位数不为18".format(sheet_id))
            if len(id_number) < 2:
                print("警告：编号为{}的表中身份证号位数小于2".format(sheet_id))
            else:
                if id_number[-2].isdigit():
                    if int(id_number[-2]) % 2 == 0:
                        gender = '女'
                    else:
                        gender = '男'

        warrant_sheet2.cell(warrant_sheet2_row_index, 5).value = member_name  # 姓名
        warrant_sheet2.cell(warrant_sheet2_row_index, 6).value = gender  # 性别
        warrant_sheet2.cell(warrant_sheet2_row_index, 7).value = id_number  # 身份证号码
        warrant_sheet2.cell(warrant_sheet2_row_index, 8).value = confirm_sheet.cell(confirm_row_index, 3).value  # 与户主关系

        warrant_sheet3.cell(warrant_sheet3_row_index, 19).value = member_name
        warrant_sheet3.cell(warrant_sheet3_row_index, 20).value = 10
        warrant_sheet3.cell(warrant_sheet3_row_index, 23).value = '成员股'

        confirm_row_index = confirm_row_index + 1
        warrant_sheet2_row_index = warrant_sheet2_row_index + 1
        warrant_sheet3_row_index = warrant_sheet3_row_index + 1
        count = count + 1
        # '家庭' not in confirm_sheet.cell(row_index, 1).value) and (confirm_sheet.cell(row_index, 3).value is not None

        if count > 14:
            print('警告：编号为{}的股权证中人数超过14，需手动调整！'.format(sheet_id))

    warrant_sheet3['T38'] = count * 10

    last4 = str(index + 1).zfill(4)

    warrant_sheet1['X22'] = para_list[0]
    warrant_sheet1['V31'] = para_list[1]
    warrant_sheet1['W52'] = para_list[2]
    warrant_sheet1['W78'] = para_list[3] + last4
    warrant_sheet1['X86'] = para_list[4]
    warrant_sheet1['AA86'] = para_list[5]
    warrant_sheet1['AB86'] = para_list[6]

    warrant_sheet3['T41'] = para_list[7]

    filename_save = last4 + confirm_sheet['A10'].value + '股权证.xlsx'
    (Path("warrant") / para_list[1]).mkdir(parents=True, exist_ok=True)

    folder_path2 = Path("warrant") / para_list[1]
    whole_save = folder_path2 / filename_save
    warrant_book.save(whole_save)
    # print(count)


def generate_share_warrants(warrant_path, confirm_path, para_list):
    folder_path, file_name = os.path.split(warrant_path)
    # folder_path = r"C:\Users\sc\Desktop\test_openpyxl"
    # file_name = 'sample.xlsx'
    old_suffix = file_name.split('.')[-1]
    if old_suffix == 'xls':
        print("transform .xls to .xlsx")
        excel_path = xls_to_xlsx(folder_path, file_name)
    elif old_suffix == 'xlsx':
        print("no need to transform file type")
        excel_path = os.sep.join([folder_path, file_name])
        # print(excel_path)
    else:
        print("wrong file type: " + old_suffix)
        return

    folder_path2, file_name2 = os.path.split(confirm_path)
    # folder_path = r"C:\Users\sc\Desktop\test_openpyxl"
    # file_name = 'sample.xlsx'
    old_suffix2 = file_name2.split('.')[-1]
    if old_suffix2 == 'xls':
        print("transform .xls to .xlsx")
        excel_path2 = xls_to_xlsx(folder_path2, file_name2)
    elif old_suffix2 == 'xlsx':
        print("no need to transform file type")
        excel_path2 = os.sep.join([folder_path2, file_name2])
        # print(excel_path)
    else:
        print("wrong file type: " + old_suffix2)
        return

    confirm_book = openpyxl.load_workbook(excel_path2, data_only=True)
    confirm_sheet_names = confirm_book.sheetnames

    for i in range(len(confirm_sheet_names)):
        generate_one_warrant(excel_path, confirm_book[confirm_sheet_names[i]], i, para_list)

    if old_suffix == 'xls':
        os.remove(excel_path)


if __name__ == '__main__':
    path_input = ''

    top = Tk()
    frame_file = Frame(top)

    # 变量path
    path = StringVar()
    # 输入框，标记，按键
    Label(top, text="确认表路径:").pack()
    # 输入框绑定变量path
    Entry(top, textvariable=path, width=30).pack()


    def selectPath():
        global path_input
        path_input = filedialog.askopenfilename(title='请选择确认登记表', filetypes=[('Excel', '.xls .xlsx')])
        path.set(path_input)
        print(path_input)


    Button(top, text="路径选择", command=selectPath).pack()

    # 1
    frame_credit_code = Frame(top)
    label_credit_code = Label(frame_credit_code, text="统一社会信用代码")
    label_credit_code.pack(side=LEFT)
    entry_credit_code = Entry(frame_credit_code)
    entry_credit_code.pack(side=RIGHT)
    frame_credit_code.pack()

    # 2
    frame_organization = Frame(top)
    label_organization = Label(frame_organization, text="组织名称")
    label_organization.pack(side=LEFT)
    entry_organization = Entry(frame_organization)
    entry_organization.pack(side=RIGHT)
    frame_organization.pack()

    # 3
    frame_representative = Frame(top)
    label_representative = Label(frame_representative, text="法定代表人")
    label_representative.pack(side=LEFT)
    entry_representative = Entry(frame_representative)
    entry_representative.pack(side=RIGHT)
    frame_representative.pack()

    # 4
    frame_number = Frame(top)
    label_number = Label(frame_number, text="股权证编号（除了后四位）")
    label_number.pack(side=LEFT)
    entry_number = Entry(frame_number)
    entry_number.pack(side=RIGHT)
    frame_number.pack()

    # 5, 6, 7
    frame_issue_date = Frame(top)
    Label(frame_issue_date, text="发证日期").pack(side=LEFT)
    entry_year = Entry(frame_issue_date, width=5)
    entry_year.pack(side=LEFT)
    Label(frame_issue_date, text="年").pack(side=LEFT)
    entry_month = Entry(frame_issue_date, width=5)
    entry_month.pack(side=LEFT)
    Label(frame_issue_date, text="月").pack(side=LEFT)
    entry_day = Entry(frame_issue_date, width=5)
    entry_day.pack(side=LEFT)
    Label(frame_issue_date, text="日").pack(side=LEFT)

    frame_issue_date.pack()

    # 8
    frame_register_date = Frame(top)
    Label(frame_register_date, text="登记日期").pack(side=LEFT)
    entry_register_date = Entry(frame_register_date)
    entry_register_date.pack(side=LEFT)
    frame_register_date.pack()


    def btn_generator_CallBack():
        para_list = [entry_credit_code.get(),
                     entry_organization.get(),
                     entry_representative.get(),
                     entry_number.get(),
                     entry_year.get(),
                     entry_month.get(),
                     entry_day.get(),
                     entry_register_date.get()]

        # messagebox.showinfo("Hello Python", para_list)

        print(os.getcwd())
        sample_path = os.path.join(os.getcwd(), 'sample')

        global path_input
        generate_share_warrants(os.path.join(sample_path, 'sample.xlsx'), path_input, para_list)

        print('股权证生成完毕。')

    B = Button(top, text="生成股权证", command=btn_generator_CallBack)
    B.pack()
    top.mainloop()
