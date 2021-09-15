import os
import winreg
import openpyxl

def get_desktop():
    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                         r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders')
    return winreg.QueryValueEx(key, "Desktop")[0]


def findNumber(in_filename):
    wb = openpyxl.load_workbook(in_filename, data_only=True)
    ws = wb.active

    nrow = ws.max_row

    for i in range(1 + 1, nrow + 1):
        if ws.cell(i, 1).value == '合计':
            #print(ws.cell(i, 2).value)
            #print(ws.cell(i, 3).value)
            global  totalF, totalP
            totalF += ws.cell(i, 2).value
            totalP += ws.cell(i, 3).value

totalF = 0
totalP = 0
file_dir = os.path.join(get_desktop(), 'to_merge')
#print(file_dir)
file_list = os.listdir(file_dir)
for file in file_list:
    # print(file)
    findNumber(os.path.join(file_dir, file))

print('户数: ', totalF)
print('人数: ', totalP)

input('Press enter to quit program.')