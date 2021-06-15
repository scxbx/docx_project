import math
import os
import win32com.client
import openpyxl


def xls_to_xlsx(folder_path, file_name):
    """
    excel  .xls 后缀 改成 .xlsx 后缀
    folder_path 文件夹路径
    file_name 文件名字 带后缀 比如 aa.xls
    """
    folder_path = folder_path.replace('/', '\\')
    file_name = file_name.replace('/', '\\')
    name, suffix = file_name.split('.')
    excel_file_path = os.path.join(folder_path, file_name)

    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')  # 要看MIME手册
    wb = excel.Workbooks.Open(excel_file_path)
    suffix = f".{suffix}x"
    new_file_name = f"{name}{suffix}"
    new_excel_file_path = os.sep.join([folder_path, new_file_name])
    # tset
    print("new_excel_file_path: " + new_excel_file_path)
    wb.SaveAs(new_excel_file_path, FileFormat=51)
    wb.Close()
    excel.Application.Quit()
    return new_excel_file_path


def replace_name(filename, summary_path):
    num_per_sheet = 54

    print("filename: " + filename)
    folder_path, file_name = os.path.split(filename)
    # folder_path = r"C:\Users\sc\Desktop\test_openpyxl"
    # file_name = 'sample.xlsx'
    old_suffix = file_name.split('.')[-1]
    if old_suffix == 'xls':
        print("transform .xls to .xlsx")
        excel_path = xls_to_xlsx(folder_path, file_name)
    elif old_suffix == 'xlsx':
        print("no need to transform file type")
        excel_path = os.sep.join([folder_path, file_name])
        # print(excel_path)
    else:
        print("wrong file type: " + old_suffix)
        return

    folder_path2, file_name2 = os.path.split(summary_path)
    # folder_path = r"C:\Users\sc\Desktop\test_openpyxl"
    # file_name = 'sample.xlsx'
    old_suffix2 = file_name2.split('.')[-1]
    if old_suffix2 == 'xls':
        print("transform .xls to .xlsx")
        excel_path2 = xls_to_xlsx(folder_path2, file_name2)
    elif old_suffix2 == 'xlsx':
        print("no need to transform file type")
        excel_path2 = os.sep.join([folder_path2, file_name2])
        # print(excel_path)
    else:
        print("wrong file type: " + old_suffix2)
        return

    book_in = openpyxl.load_workbook(excel_path, data_only=True)

    # sheet_in = book_in.active

    summary_book = openpyxl.load_workbook(excel_path2, data_only=True)
    summary_sheet = summary_book.active

    for i in range(1, 12):
        if summary_sheet.cell(2, i).value is not None and type(
                summary_sheet.cell(2, i).value) == str and '集体经济组织名称' in summary_sheet.cell(2, i).value:
            # print(summary_sheet.cell(2, i).value)
            completeName = summary_sheet.cell(2, i).value.split('：')[1]
            # print(completeName)
            # 县
            county = completeName.split('县', 1)[0] + '县'
            county_right = completeName.split('县', 1)[1]
            # (county)
            # 镇 乡
            if '镇' in county_right:
                town = county_right.split('镇', 1)[0] + '镇'
                town_right = county_right.split('镇', 1)[1]
            elif '乡' in county_right:
                town = county_right.split('乡', 1)[0] + '乡'
                town_right = county_right.split('乡', 1)[1]
            # print(town)
            # 村
            if '社区' in town_right:
                village = town_right.split('社区', 1)[0] + '社区'
                village_right = town_right.split('社区', 1)[1]
            elif '居委会' in town_right:
                village = town_right.split('居委会', 1)[0] + '居委会'
                village_right = town_right.split('居委会', 1)[1]
            elif '村' in town_right:
                village = town_right.split('村', 1)[0] + '村'
                village_right = town_right.split('村', 1)[1]
            else:
                print('找不到社区、居委会、村')
                village = '找不到社区、居委会、村'
                village_right = town_right
            # print(village)
            # 组
            if '股份经济合作联合社' in village_right:
                group = ''
                isCommittee = True
            else:
                group = village_right.split('组', 1)[0]
                isCommittee = False
            # print(group)
        else:
            print('{} cannot find'.format(i))

    for sheet_in in book_in.worksheets:
        # print(type(sheet_in.title))
        if isCommittee is False:
            sheet_in.title = sheet_in.title.replace("东排湾村民小组", group + '村民小组')
        else:
            if '村' in village:
                sheet_in.title = sheet_in.title.replace("东排湾村民小组", village + '委会')
            else:
                sheet_in.title = sheet_in.title.replace("东排湾村民小组", village)

        for row in sheet_in.iter_rows():
            for cell in row:
                # print(cell.value, end=" ")

                if type(cell.value) is str:
                    cell.value = cell.value.replace('澄迈县', county)
                    cell.value = cell.value.replace("加乐镇", town)
                    if isCommittee == False:
                        if '村' in village:
                            cell.value = cell.value.replace("加桐村委会东排湾村民小组", village + '委会' + group + '村民小组')
                            # (village + '委会' + group + '村民小组')
                        else:
                            cell.value = cell.value.replace("加桐村委会东排湾村民小组", village + group + '村民小组')
                            # print(village + group + '村民小组')
                        cell.value = cell.value.replace("加桐村委会东排湾组", village + group + '组')
                        # print(village + group + '组')
                        cell.value = cell.value.replace("加桐村东排湾组", village + group + '组')
                        # print(village + group + '组')
                    else:
                        if '村' in village:
                            cell.value = cell.value.replace("加桐村委会东排湾村民小组", village + '委会')
                        else:
                            cell.value = cell.value.replace("加桐村委会东排湾村民小组", village)
                        cell.value = cell.value.replace("加桐村委会东排湾组", village)
                        cell.value = cell.value.replace("加桐村东排湾组", village)
                        cell.value = cell.value.replace("股份经济合作社", "股份经济合作联合社")

    sheets_in = book_in.sheetnames
    # sheet_register = book_in[sheets_in[3]]

    row_count = 0
    for row in summary_sheet:
        if not all([cell.value == None for cell in row]):
            row_count += 1
    print("row_count: {}".format(row_count))
    householder_list = []
    for i in range(4, row_count):
        if summary_sheet.cell(i, 2).value is not None:
            householder_list.append(summary_sheet.cell(i, 2).value)

        # if summary_sheet.cell(2, i).value is not None:
        # print(summary_sheet.cell(2, i).value)

    # num_new_sheet = int((len(householder_list) - num_per_sheet) / (num_per_sheet + 1)) + 1
    num_new_sheet = math.ceil(len(householder_list) / num_per_sheet) - 1

    for ii in range(num_new_sheet):
        new_sheet = book_in.copy_worksheet(book_in[sheets_in[3]])
        # new_sheet.title = sheets_in[3], "(卷," + str(ii + 2) + ")"
        new_sheet.title = "{}(卷{}）".format(sheets_in[3], str(ii + 2))
        # new_sheet.cell(4, 4).value = new_sheet.cell(5, 4).value
        # new_sheet.delete_rows(num_per_sheet + 4)

    new_sheets_in = book_in.sheetnames

    for j in range(len(householder_list)):

        sheet_index = 3 + int(j / num_per_sheet)

        # print(j)
        # print(sheet_index)
        if type(book_in[new_sheets_in[sheet_index]].cell(j % num_per_sheet + 5, 4).value) is str:
            if type(householder_list[j]) == int:
                print("householder_list[{}]: {} is int which should be str".format(j, householder_list[j]))
            book_in[new_sheets_in[sheet_index]].cell(j % num_per_sheet + 5, 4).value = book_in[
                new_sheets_in[sheet_index]].cell(j % num_per_sheet
                                                 + 5,
                                                 4).value.replace(
                "householders", householder_list[j])
            # book_in[new_sheets_in[sheet_index]].cell(j % num_per_sheet + 4, 1).value = j % num_per_sheet + 1

    for i in range(num_new_sheet):
        book_in[new_sheets_in[4 + i]].delete_rows(4)
        for j in range(num_per_sheet):
            book_in[new_sheets_in[4 + i]].cell(j + 4, 1).value = j + 1

    for i in range(num_per_sheet):
        if type(book_in[new_sheets_in[-1]].cell(i + 4, 4).value) == str and 'householders' in book_in[
            new_sheets_in[-1]].cell(i + 4, 4).value:
            # print('householders')
            len_delete = num_per_sheet - len(householder_list) % num_per_sheet
            book_in[new_sheets_in[-1]].delete_rows(i + 4, len_delete)

    filename_save = (os.path.split(summary_path))[1].split('.')[0] + '目录.xlsx'
    folder_path2 = folder_path + r'\..\menu'
    whole_save = os.path.join(folder_path2, filename_save)
    book_in.save(whole_save)

    if old_suffix == 'xls':
        os.remove(excel_path)
    if old_suffix2 == 'xls':
        os.remove(excel_path2)
    return


if __name__ == '__main__':

    print(os.getcwd())
    path = os.path.join(os.getcwd(), 'summary')
    sample_path = os.path.join(os.getcwd(), 'sample')
    for filename_in in os.listdir(path):
        print(os.path.join(path, filename_in))
        # read_summary(os.path.join(path, filename_in))
        replace_name(os.path.join(sample_path, 'sample.xlsx'), os.path.join(path, filename_in))
    input('Press any key to quit program.')
