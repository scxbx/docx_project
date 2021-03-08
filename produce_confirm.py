import os

import openpyxl
import win32com.client


def xls_to_xlsx(folder_path, file_name):
    """
    excel  .xls 后缀 改成 .xlsx 后缀
    folder_path 文件夹路径
    file_name 文件名字 带后缀 比如 aa.xls
    """
    folder_path = folder_path.replace('/', '\\')
    file_name = file_name.replace('/', '\\')
    name, suffix = file_name.split('.')
    excel_file_path = os.path.join(folder_path, file_name)

    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')  # 要看MIME手册
    wb = excel.Workbooks.Open(excel_file_path)
    suffix = f".{suffix}x"
    new_file_name = f"{name}{suffix}"
    new_excel_file_path = os.sep.join([folder_path, new_file_name])
    # tset
    print("new_excel_file_path: " + new_excel_file_path)
    wb.SaveAs(new_excel_file_path, FileFormat=51)
    wb.Close()
    excel.Application.Quit()
    return new_excel_file_path

def produce_a_confirm(filename, start_row):
    # 数据开始的行数
    start_row = start_row
    folder_path, file_name = os.path.split(filename)

    wb = openpyxl.load_workbook(filename)
    sheet_in = wb.active

    # worksheet.merged_cells获取已经合并单元格的信息；再使用worksheet.unmerge_cells()拆分单元格；
    m_list = sheet_in.merged_cells
    cr = []
    for m_area in m_list:
        # 合并单元格的起始行坐标、终止行坐标。。。。，
        r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
        # 纵向合并单元格的位置信息提取出
        if r2 - r1 > 0:
            cr.append((r1, r2, c1, c2))

    for r in cr:
        sheet_in.unmerge_cells(start_row=r[0], end_row=r[1], start_column=r[2], end_column=r[3])


    book_out = openpyxl.load_workbook(os.path.join(folder_path, r'..\copy.xlsx'))
    nrow = sheet_in.max_row
    print ('nrow: ' + str(nrow))

    flag = 1
    num_family = 0
    while flag < nrow + 1:
        if (sheet_in.cell(flag, 1).value == '合计' or sheet_in.cell(flag, 1).value == '汇总'):
            num_family = sheet_in.cell(flag, 2).value
        flag += 1
    print(num_family)
    '''
    if isinstance(sheet_in.cell(nrow, 2).value, int):
        num_family = sheet_in.cell(nrow, 2).value
    else:
        num_family = 500
    '''
    for i in range(num_family):
        book_out.copy_worksheet(book_out.active)
    sheets_out = book_out.sheetnames



    serial_number = 0
    # col_num_list = [1, 2, 3, 7, 8, 9]
    current_merge_head = start_row
    # row_now = start_row
    count_in_family = 0
    sheet_out = book_out[sheets_out[serial_number]]
    for i in range(start_row, flag - 2):
        if sheet_in.cell(i, 1).value is not None:
            print(serial_number)
            print(sheets_out[serial_number])
            sheet_out = book_out[sheets_out[serial_number]]
            # 序号
            sheet_out['K2'] = serial_number + 1
            # 户主
            sheet_out['C3'] = sheet_in.cell(i, 2).value
            # 集体经济组织名称 地址
            sheet_out['C2'] = sheet_in.cell(start_row, 9).value.strip()
            sheet_out['C4'] = sheet_in.cell(start_row, 9).value.strip()
            # 电话
            sheet_out['H3'] = sheet_in.cell(i, 10).value
            # 家庭成员总数
            sheet_out['K8'] = '共 {} 人'.format(sheet_in.cell(i, 3).value)
            # 邮政编码
            sheet_out['J4'] = 572900

            serial_number += 1
            count_in_family = 0

        # 姓名
        sheet_out.cell(10 + count_in_family, 1).value = sheet_in.cell(i, 4).value
        # 关系
        sheet_out.cell(10 + count_in_family, 3).value = sheet_in.cell(i, 5).value
        # 身份证
        sheet_out.cell(10 + count_in_family, 5).value = sheet_in.cell(i, 7).value
        if sheet_in.cell(i, 5).value in ['户主', '本人']:
            sheet_out['H5'] = sheet_in.cell(i, 7).value
        # 备注
        sheet_out.cell(10 + count_in_family, 9).value = sheet_in.cell(i, 11).value
        count_in_family += 1

    filename_save = (os.path.split(filename))[1].split('.')[0] + '_确认登记表.xlsx'
    folder_path2 = folder_path + r'\..\confirm'
    whole_save = os.path.join(folder_path2, filename_save)


    # 修改sheetname
    for i in range(len(sheets_out)):
        ws = book_out[sheets_out[i]]
        ws.title = str(i + 1)
    book_out.save(whole_save)
    book_out.close()

if __name__ == '__main__':
    path = os.path.join(os.getcwd(), 'summary')
    for filename_in in os.listdir(path):
        # print(os.path.join(path, filename))
        print(filename_in)
        produce_a_confirm(os.path.join(path, filename_in), 4)

    input('Press any key to quit program.')
