# *-* coding-utf-8 *-*
import sys


import os

def merge_Excel(excelpath_list, new_excel_path=os.path.dirname(__file__)+'\\合并的文件.xlsx', source_all_sheet=True):
    '''
    将多个excel文件，合并到一个新的excel里。
    :param excelpath_list: 要合并的excel文件路径列表
    :param new_excel_path: 合并后的excel文件路径
    :param source_all_sheet: 是否合并源Excel中的所有工作表
    :return:
    '''
    if len(excelpath_list) == 0:
        return False
    from openpyxl import Workbook
    from openpyxl import load_workbook

    sheet_count = 0

    # 实例化一个新的工作簿，用户保存合并的数据
    wb_new = Workbook()

    for excel_path in excelpath_list:  # 第1层循环，遍历所有的Excel文件 =========================================
        wb = load_workbook(excel_path)  # 载入源excel文件，获取工作簿
        wb.guess_types = True  # 猜测格式类型
        # 获取 工作簿中的所有sheet
        if source_all_sheet:
            ws_all = wb.worksheets
        else:
            ws_all = [wb.active]

        for sheet in ws_all:  # 第2层循环，遍历每个Excel的sheet =========================================
            # 获取 sheet中的有效行数、列数
            count_row = 0  # 工作表的有效行数
            count_col = 0  # 工作表的有效列数
            for row in sheet.rows:
                count_row += 1
            for col in sheet.columns:
                count_col += 1
            if count_row==0 or count_col==0: #如果有效行数/列数为0，代表当前表没有数据 [重要]
                continue

            # 读取sheet的内容，写入到新工作簿的工作表中
            list_all = []
            row_range = sheet[1:count_row]
            list_row = []
            tag=False

            for row in row_range:
                if  type(row) is not tuple:   # 判断row的类型是不是元祖，如果不是元祖，代表当前sheet只有一行数据，row就是单元格对象; 反之，row是包含了整行单元格对象的元祖
                    tag = True
                    break
            if tag:  # 代表当前sheet只有一行数据，row就是单元格对象;
                for cell in row_range:  # 第3层循环  遍历工作表的行、列。封装数据  ==============================
                     list_row.append(cell.value)
                list_all.append(list_row)  # 将1行中的所有列的数据(列表类型)，在添加进总列表里
            else: #  代表   当前sheet有多行数据
                for row in row_range:
                    list_row = []  # 清空
                    for cell in row:
                        list_row.append(cell.value)  # 遍历1行中的每列数据，读取后，添加进一个列表
                    list_all.append(list_row)  # 将1行中的所有列的数据(列表类型)，在添加进总列表里

            # 将源sheet中提取后，封装的数据写入新的工作表
            ws_temp = wb_new.create_sheet('sheet-%s' % sheet_count)
            for item in list_all:  # 第3层循环  将提取的原sheet数据 循环写入到新的sheet里
                # 判断用户选择是否合并到一个sheet
                ws_temp.append(item)
            sheet_count += 1

    # 保存文件
    wb_new.save(new_excel_path)
    return True

def run():
    # 2. 组织 合成后的新excel文件的绝对路径，判断该文件，是否已经存在
    new_excel_path = 'merged.xlsx'
    folder_path = 'C:/Users/sc/Desktop/to_merge'
    if os.path.exists(new_excel_path):
        os.remove(new_excel_path)  # 删除该文件

    # 3.组织所有要合并的excel文件的绝对路径，添加进list
    list_file=os.listdir(folder_path) # 获取日期文件夹下的所有文件名称列表
    list_file_path=[ ]
    for item in list_file:
        list_file_path.append(os.path.join(folder_path,item) )
    print(list_file_path)

    # 4.调用函数，合并 所有excel
    res = merge_Excel(list_file_path,new_excel_path)
    if res:
        print("合并成功")


if __name__ == "__main__":
    run()