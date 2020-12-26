import os
import win32com.client

import openpyxl


def xls_to_xlsx(folder_path, file_name):
    """
    excel  .xls 后缀 改成 .xlsx 后缀
    folder_path 文件夹路径
    file_name 文件名字 带后缀 比如 aa.xls
    """
    folder_path = folder_path.replace('/', '\\')
    file_name = file_name.replace('/', '\\')
    name, suffix = file_name.split('.')
    excel_file_path = os.path.join(folder_path, file_name)

    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')  # 要看MIME手册
    wb = excel.Workbooks.Open(excel_file_path)
    suffix = f".{suffix}x"
    new_file_name = f"{name}{suffix}"
    new_excel_file_path = os.sep.join([folder_path, new_file_name])
    # tset
    print("new_excel_file_path: " + new_excel_file_path)
    wb.SaveAs(new_excel_file_path, FileFormat=51)
    wb.Close()
    excel.Application.Quit()
    return new_excel_file_path


def read_summary(filename):
    print("filename: " + filename)
    folder_path, file_name = os.path.split(filename)
    # folder_path = r"C:\Users\sc\Desktop\test_openpyxl"
    # file_name = 'sample.xlsx'
    old_suffix = file_name.split('.')[-1]
    if old_suffix == 'xls':
        print("transform .xls to .xlsx")
        excel_path = xls_to_xlsx(folder_path, file_name)
    elif old_suffix == 'xlsx':
        print("no need to transform file type")
        excel_path = os.sep.join([folder_path, file_name])
        # print(excel_path)
    else:
        print("wrong file type: " + old_suffix)
        return

    book_in = openpyxl.load_workbook(excel_path, data_only=True)

    sheet_in = book_in.active

    if old_suffix == 'xls':
        os.remove(excel_path)

    m_list = sheet_in.merged_cells  # 合并单元格的位置信息，可迭代对象（单个是一个'openpyxl.worksheet.cell_range.CellRange'对象），print后就是excel坐标信息
    merged_cells_rows = []
    for m_area in m_list:
        # 合并单元格的起始行坐标、终止行坐标。。。。，
        r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
        # 纵向合并单元格的位置信息提取出
        if r2 - r1 > 0:
            merged_cells_rows.append((r1, r2))
            # print('符合条件%s' % str(m_area))

    # print(merged_cells_rows)
    no_dupl_merged_rows = list(set(merged_cells_rows))
    # print('合计', len(no_dupl_merged_rows))

    #sorted_cr = sorted(no_dupl_merged_rows, key=lambda x: x[1])
    no_dupl_merged_rows.sort()
    # print(no_dupl_merged_rows)
    # print(len(no_dupl_merged_rows))
    to_add_list = []
    for i in range(len(no_dupl_merged_rows)):
        right = no_dupl_merged_rows[i][0]
        left = no_dupl_merged_rows[i - 1][1]
        dif = right - left
        if i > 0 and dif > 1:
            for j in range(1, dif):
                to_add_list.append((left+j, left+j))

    # print(to_add_list)

    final_rows = no_dupl_merged_rows + to_add_list

    final_rows.sort()
    # print(final_rows)

    headcounts = []
    for final_row in final_rows:
        headcounts.append(final_row[1] - final_row[0] + 1)

    # 老城 在headcounts删去第一个合并单元格
    del headcounts[0]
    # print('headcounts: ', headcounts)
    # print('len of headcounts: ', len(headcounts))

    book_out = openpyxl.load_workbook(os.path.join(folder_path, r'..\copy.xlsx'))
    for i in range(len(headcounts) - 1):
        book_out.copy_worksheet(book_out.active)
    sheets_out = book_out.sheetnames
    '''
    row_now = 4
    '''
    row_now = 4
    for i in range(len(headcounts)):
        if i < len(sheets_out):
            # fill in one sheet_in (one family)
            sheet_out = book_out[sheets_out[i]]


            '''
            # -------------------陵水 第四行开始 ---------------------------
            # 序号
            sheet_out['K2'] = i + 1
            # 户主
            sheet_out['C3'] = sheet_in.cell(row_now, 2).value
            # 集体经济组织名称 地址
            sheet_out['C2'] = sheet_in.cell(2, 3).value
            sheet_out['C4'] = sheet_in.cell(2, 3).value
            # 电话
            sheet_out['H3'] = sheet_in.cell(row_now, 10).value
            # 家庭成员总数
            sheet_out['K8'] = '共 {} 人'.format(headcounts[i])
            # 邮政编码
            sheet_out['J4'] = 572435
            
                        #unmerge
            # worksheet.merged_cells获取已经合并单元格的信息；再使用worksheet.unmerge_cells()拆分单元格；
            m_list = sheet_out.merged_cells
            cr = []
            for m_area in m_list:
                # 合并单元格的起始行坐标、终止行坐标。。。。，
                r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
                # 纵向合并单元格的位置信息提取出
                if r2 - r1 > 0:
                    cr.append((r1, r2, c1, c2))

            for r in cr:
                sheet_out.unmerge_cells(start_row=r[0], end_row=r[1], start_column=r[2], end_column=r[3])


            # 家庭成员信息
            for j in range(headcounts[i]):
                # 姓名
                sheet_out.cell(10 + j, 1).value = sheet_in.cell(row_now, 4).value
                # 关系
                # sheet_out.unmerge_cells(start_row=10 + j, end_row=10 + j, start_column=3, end_column=4)
                sheet_out.cell(10 + j, 3).value = sheet_in.cell(row_now, 5).value
                # sheet_out.merge_cells(start_row=10 + j, end_row=10 + j, start_column=3, end_column=4)
                # 身份证
                # sheet_out.unmerge_cells(start_row=10 + j, end_row=10 + j, start_column=5, end_column=8)
                sheet_out.cell(10 + j, 5).value = sheet_in.cell(row_now, 7).value
                if sheet_in.cell(row_now, 5).value in ['户主','本人']:
                    sheet_out['H5'] = sheet_in.cell(row_now, 7).value
                # sheet_out.merge_cells(start_row=10 + j, end_row=10 + j, start_column=5, end_column=8)
                # 备注
                # sheet_out.unmerge_cells(start_row=10 + j, end_row=10 + j, start_column=9, end_column=11)
                sheet_out.cell(10 + j, 9).value = sheet_in.cell(row_now, 11).value
                # sheet_out.merge_cells(start_row=10 + j, end_row=10 + j, start_column=9, end_column=11)


                row_now += 1

            for r in cr:
                # worksheet.merge_cells()合并单元格
                sheet_out.merge_cells(start_row=r[0], end_row=r[1], start_column=r[2], end_column=r[3])


            '''
            # -------------------老城 第六行开始 无证件类型---------------------------
            # 序号
            sheet_out['K2'] = i + 1
            # 户主
            sheet_out['C3'] = sheet_in.cell(row_now, 2).value
            # 集体经济组织名称 地址
            sheet_out['C2'] = sheet_in.cell(2, 3).value
            sheet_out['C4'] = sheet_in.cell(2, 3).value
            # 电话
            sheet_out['H3'] = sheet_in.cell(row_now, 10).value
            # 家庭成员总数
            sheet_out['K8'] = '共 {} 人'.format(headcounts[i])
            # 邮政编码
            sheet_out['J4'] = 572435




            #unmerge
            # worksheet.merged_cells获取已经合并单元格的信息；再使用worksheet.unmerge_cells()拆分单元格；
            m_list = sheet_out.merged_cells
            cr = []
            for m_area in m_list:
                # 合并单元格的起始行坐标、终止行坐标。。。。，
                r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
                # 纵向合并单元格的位置信息提取出
                if r2 - r1 > 0:
                    cr.append((r1, r2, c1, c2))

            for r in cr:
                sheet_out.unmerge_cells(start_row=r[0], end_row=r[1], start_column=r[2], end_column=r[3])


            # 家庭成员信息
            for j in range(headcounts[i]):
                # 姓名
                sheet_out.cell(10 + j, 1).value = sheet_in.cell(row_now, 4).value
                # 关系
                # sheet_out.unmerge_cells(start_row=10 + j, end_row=10 + j, start_column=3, end_column=4)
                sheet_out.cell(10 + j, 3).value = sheet_in.cell(row_now, 5).value
                # sheet_out.merge_cells(start_row=10 + j, end_row=10 + j, start_column=3, end_column=4)
                # 身份证
                # sheet_out.unmerge_cells(start_row=10 + j, end_row=10 + j, start_column=5, end_column=8)
                sheet_out.cell(10 + j, 5).value = sheet_in.cell(row_now, 7).value
                if sheet_in.cell(row_now, 5).value in ['户主','本人']:
                    sheet_out['H5'] = sheet_in.cell(row_now, 7).value
                # sheet_out.merge_cells(start_row=10 + j, end_row=10 + j, start_column=5, end_column=8)
                # 备注
                # sheet_out.unmerge_cells(start_row=10 + j, end_row=10 + j, start_column=9, end_column=11)
                sheet_out.cell(10 + j, 9).value = sheet_in.cell(row_now, 11).value
                # sheet_out.merge_cells(start_row=10 + j, end_row=10 + j, start_column=9, end_column=11)


                row_now += 1

            for r in cr:
                # worksheet.merge_cells()合并单元格
                sheet_out.merge_cells(start_row=r[0], end_row=r[1], start_column=r[2], end_column=r[3])



    # print(sheet_in['D15'].value)

    # print(book_out[sheets_out[i]].merged_cells)
    filename_save = (os.path.split(filename))[1].split('.')[0] + '确认表.xlsx'
    folder_path2 = folder_path + r'\..\confirm'
    whole_save = os.path.join(folder_path2, filename_save)

    book_out.save(whole_save)

if __name__ == '__main__':
    # read_summary(r'C:\Users\sc\Desktop\test_docx\summary_sample.xlsx')
    # print(os.getcwd())
    path = os.path.join(os.getcwd(), 'summary')
    for filename_in in os.listdir(path):
        # print(os.path.join(path, filename))
        read_summary(os.path.join(path, filename_in))

    input('Press any key to quit program.')