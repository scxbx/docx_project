import os
import re
from tkinter import Tk, filedialog, Button
import openpyxl
from openpyxl.styles import PatternFill


def split_name(organization, _xiang_or_zhen):
    level1 = '县'
    level2 = _xiang_or_zhen

    re_str = '{}|{}'.format(level1, level2)
    result = re.split(re_str, organization)
    print(result)
    return result


def xiang_or_zhen(organization):
    if '乡' in organization:
        return '乡'
    elif '镇' in organization:
        return '镇'
    else:
        print("既不是乡也不是镇")
        return ''


def modify_menu(filename_input, sample):
    folder_path_in, file_name_in = os.path.split(filename_input)
    book_in = openpyxl.load_workbook(filename_input, data_only=True)
    sheet_in = book_in.worksheets[0]

    book_sample = openpyxl.load_workbook(sample, data_only=True)
    sheet_sample = book_sample.worksheets[0]

    # -----------------------------------------------------------------------------

    org = sheet_in['C4'].value
    x_or_z = xiang_or_zhen(org)
    split_result = split_name(org, x_or_z)

    for i in range(4, 20):
        sheet_in.cell(i, 3).value = sheet_sample.cell(i, 3).value
        sheet_in.cell(i, 4).value = sheet_sample.cell(i, 4).value

    for j in range(20, 24):
        sheet_in.delete_rows(20)  # 表示删除表格的第j行

    # 替换组织名称
    for row in sheet_in.iter_rows():
        for cell in row:
            if type(cell.value) is str:
                cell.value = cell.value.replace('陵水黎族自治县', '{}县'.format(split_result[0]))
                cell.value = cell.value.replace('提蒙乡', '{}{}'.format(split_result[1], x_or_z))
                cell.value = cell.value.replace('曾山村委会曾山一村民小组', '{}'.format(split_result[2]))

    # 主题颜色
    sheet2_in = book_in.worksheets[1]
    for row in sheet2_in.iter_rows():
        for cell in row:
            cell.fill = PatternFill('solid', fgColor='ffffff')

    # filename_save = (os.path.split(filename_in))[1].split('.')[0] + '.xlsx'
    # folder_path2 = folder_path_in + r'\..\new menu'
    # whole_save = os.path.join(folder_path2, filename_save)
    book_in.save(filename_input)

    return


def clear_one_cell(filename_input):
    book_in = openpyxl.load_workbook(filename_input, data_only=True)
    if len(book_in.worksheets) <= 3:
        print('{} has only {} sheets.'.format(os.path.split(filename_input)[1], len(book_in.worksheets)))
        return
    sheet4_in = book_in.worksheets[3]
    print(sheet4_in['D4'])
    sheet4_in['D4'] = ''
    book_in.save(filename_input)
    return


def print_sheet1_once(filename_input, count):
    book_in = openpyxl.load_workbook(filename_input, data_only=True)
    sheet1_name = book_in.sheetnames[0]
    is_that = '村委会成员界定' in sheet1_name \
              or '村民委员会成员界定' in sheet1_name \
              or '社区成员界定' in sheet1_name \
              or '居委会成员界定' in sheet1_name \
              or '居民委员会成员界定' in sheet1_name
    if is_that:
        print('{}\tsheet1: {}  isContain委会成员界定: {}'.format(count, sheet1_name, is_that))
        for sheet_name in book_in.sheetnames:
            if '户口本复印件NCG03-4'in sheet_name:
                print("Delete {}".format(sheet_name))
                del book_in[sheet_name]
    book_in.save(filename_input)
    return


def num_of_people_once(filename_input):
    book_in = openpyxl.load_workbook(filename_input, data_only=True)
    sheet1 = book_in.worksheets[0]
    for row in range(sheet1.max_row):
        if sheet1.cell(row + 1, 1) is not None and '合计' == sheet1.cell(row + 1, 1).value:
            print('{} {}'.format(os.path.split(filename_input)[1], sheet1.cell(row + 1, 3).value))


if __name__ == '__main__':
    def selectPathAndMod():
        cur = filedialog.askopenfilenames(
            filetypes=[('excel', ('.xlsx', '.xls'))])
        if cur:
            for filename_in in cur:
                # print(cur)
                # read_summary(os.path.join(path, filename_in))
                modify_menu(filename_in, os.path.join(sample_path, 'sample.xlsx'))
            print('----------------------------------------------------------')


    def traverse():
        folder_selected = filedialog.askdirectory()
        for root, dirs, files in os.walk(folder_selected):
            for file in files:
                if file.endswith('.xlsx'):
                    modify_menu(os.path.join(root, file), os.path.join(sample_path, 'sample.xlsx'))
            print('----------------------------------------------------------')


    def clear_traverse():
        folder_selected = filedialog.askdirectory()
        for root, dirs, files in os.walk(folder_selected):
            for file in files:
                if file.endswith('.xlsx'):
                    clear_one_cell(os.path.join(root, file))
            print('----------------------------------------------------------')





    def print_sheet1():
        count = 1
        folder_selected = filedialog.askdirectory()
        for root, dirs, files in os.walk(folder_selected):
            for file in files:
                if file.endswith('.xlsx'):
                    print_sheet1_once(os.path.join(root, file), count)
            count += 1
            print('----------------------------------------------------------')





    def num_of_people():
        folder_selected = filedialog.askdirectory()
        for root, dirs, files in os.walk(folder_selected):
            for file in files:
                if file.endswith('.xlsx'):
                    num_of_people_once(os.path.join(root, file))

    sample_path = os.path.join(os.getcwd(), 'sample')

    root = Tk(className='Modify Menu')


    Button(root, text="选择文件（可多个）", command=selectPathAndMod, width=20).pack()
    Button(root, text="选择路径", command=traverse, width=20).pack()
    Button(root, text="清空一格", command=clear_traverse, width=20).pack()
    Button(root, text="test", command=print_sheet1, width=20).pack()
    Button(root, text="总人数", command=num_of_people, width=20).pack()
    root.mainloop()
