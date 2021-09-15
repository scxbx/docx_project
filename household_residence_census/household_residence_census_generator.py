import os

import openpyxl

from menu_generator import xls_to_xlsx
from registration_form_for_share_certificate_application.registration_share import get_max_row


def gen_household_residence_census(sample_path, summary_path, save_path):
    print('-----------------------------------------开始----------------------------------------------')
    print("模板: " + sample_path)
    print('输入汇总表：' + summary_path)

    # ----------------------------------通用开头 Start------------------------------------------
    folder_path, file_name = os.path.split(sample_path)
    # folder_path = r"C:\Users\sc\Desktop\test_openpyxl"
    # file_name = 'sample.xlsx'
    old_suffix = file_name.split('.')[-1]
    if old_suffix == 'xls':
        print("transform .xls to .xlsx")
        excel_path = xls_to_xlsx(folder_path, file_name)
    elif old_suffix == 'xlsx':
        print("no need to transform file type")
        excel_path = os.sep.join([folder_path, file_name])
        # print(excel_path)
    else:
        print("wrong file type: " + old_suffix)
        return

    folder_path2, file_name2 = os.path.split(summary_path)
    # folder_path = r"C:\Users\sc\Desktop\test_openpyxl"
    # file_name = 'sample.xlsx'
    old_suffix2 = file_name2.split('.')[-1]
    if old_suffix2 == 'xls':
        print("transform .xls to .xlsx")
        excel_path2 = xls_to_xlsx(folder_path2, file_name2)
    elif old_suffix2 == 'xlsx':
        print("no need to transform file type")
        excel_path2 = os.sep.join([folder_path2, file_name2])
        # print(excel_path)
    else:
        print("wrong file type: " + old_suffix2)
        return

    book_summary = openpyxl.load_workbook(excel_path2, data_only=True)
    sheet_summary = book_summary.active
    if '总表' in book_summary.sheetnames:
        sheet_summary = book_summary['总表']
        print('Here!')

    book_sample = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_sample = book_sample.active

    # for data in sheet_in['A']:
    #    print(data.value)

    # Get the number of rows in worksheet
    max_row = get_max_row(sheet_summary)
    # print(max_row)
    # ----------------------------------通用开头 End------------------------------------------

    num_member = 0
    for row in range(3, max_row):
        serial_number = sheet_summary.cell(row + 1, 1).value  # 序号
        if serial_number in ['合计', '总计', '汇总']:
            for i in range(3):
                sheet_sample.merge_cells(start_row=row - num_member, end_row=row - 1, start_column=i + 1,
                                         end_column=i + 1)
            break

        if serial_number is not None:
            sheet_sample.cell(row, 2).value = sheet_summary.cell(row + 1, 2).value  # 户主

            if row != 3:
                for i in range(3):
                    # print(row)
                    # print('num: {}'.format(num_member))
                    sheet_sample.merge_cells(start_row=row - num_member, end_row=row - 1, start_column=i + 1,
                                             end_column=i + 1)

            num_member = int(sheet_summary.cell(row + 1, 3).value)
            sheet_sample.cell(row, 3).value = num_member  # 人数

        id_number = sheet_summary.cell(row + 1, 7).value
        sheet_sample.cell(row, 4).value = sheet_summary.cell(row + 1, 4).value  # 家庭成员
        sheet_sample.cell(row, 5).value = sheet_summary.cell(row + 1, 5).value  # 关系
        sheet_sample.cell(row, 8).value = sheet_summary.cell(row + 1, 6).value  # 性别
        sheet_sample.cell(row, 18).value = sheet_summary.cell(row + 1, 10).value  # 备注
        sheet_sample.cell(row, 13).value = id_number  # 身份证号

        if id_number is None or id_number == '' or len(id_number) != 18:
            print('第{}行缺失身份证或身份证位数不为18！'.format(row))
        else:
            sheet_sample.cell(row, 10).value = \
                "%s年%s月%s日" % (id_number[6:10], id_number[10:12], id_number[12:14])  # 出生日期
            sheet_sample.cell(row, 14).value = 2021 - int(id_number[6:10])  # 年龄

    book_sample.save(save_path)
    print('生成文件{}\n'.format(save_path))

    if old_suffix == 'xls':
        os.remove(excel_path)
    if old_suffix2 == 'xls':
        os.remove(excel_path2)
    return


if __name__ == "__main__":
    path = os.path.join(os.getcwd(), 'summary')
    sample_f = os.path.join(os.getcwd(), 'sample')
    # for filename_in in os.listdir(path):
        # read_summary(os.path.join(path, filename_in))
        #
    for filepath, dirnames, filenames in os.walk(path):
        for filename in filenames:
            long_path = filepath.replace('summary', 'result', 1)
            if not os.path.exists(long_path):
                os.makedirs(long_path)
            save_path = os.path.join(long_path, filename).split('.')[0] + '农村宅基地户籍情况调查表.xlsx'
            gen_household_residence_census(os.path.join(sample_f, 'sample.xlsx'), os.path.join(filepath, filename),
                                           save_path)
    input('生成完毕。按回车键以结束程序。')
